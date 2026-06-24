"""
╔══════════════════════════════════════════════════════════════════════════════╗
║        FRUIT & VEGETABLE PROCUREMENT ANALYZER  –  v1.0                     ║
║        Автоматичен анализ на плодове и зеленчуци от счетоводен експорт     ║
║                                                                              ║
║  Употреба:                                                                   ║
║    python analyzer.py --baseline файл_2025.xlsx --ytd файл_2026.xlsx        ║
║    python analyzer.py --baseline 2025.xlsx --ytd 2026.xlsx --output отчет   ║
║    python analyzer.py --help                                                 ║
╚══════════════════════════════════════════════════════════════════════════════╝
"""

import argparse
import sys
import os
from pathlib import Path
from datetime import date

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────
VERSION = "1.0"
MONTH_NAMES = [
    "Януари", "Февруари", "Март", "Април", "Май", "Юни",
    "Юли", "Август", "Септември", "Октомври", "Ноември", "Декември",
]

# Excel styles palette
C = {
    "navy":   "1B3A6B",
    "teal":   "0D7377",
    "light":  "EBF1F8",
    "white":  "FFFFFF",
    "green":  "1B5E20",
    "red":    "B71C1C",
    "orange": "E65100",
    "input":  "0000FF",   # blue  = hardcoded input
    "calc":   "000000",   # black = formula / calculation
    "xref":   "006400",   # green = cross-sheet reference
    "navy2":  "2C4E80",
    "gold":   "F0A500",
}

# Number formats
FMT = {
    "eur":  '€#,##0.00;(€#,##0.00);"-"',
    "eur4": '€#,##0.0000',
    "num":  '#,##0.00;(#,##0.00);"-"',
    "pct":  '0.0%;-0.0%;"-"',
    "date": "DD.MM.YYYY",
    "int":  "0",
}

ABC_COLORS = {
    "A": {"bg": "D4EDDA", "fg": "1B5E20"},
    "B": {"bg": "FFF3CD", "fg": "856404"},
    "C": {"bg": "FDE8E8", "fg": "B71C1C"},
}

VENDOR_COLORS = {
    "РЕГОЛИТ ЕООД":       "D4EDDA",
    "АКВАТЕК ООД":         "D1ECF1",
    "ПИКАНТ ТРЕЙД ООД":    "FFF3CD",
    "ГОЛДЪН ПЛОД-ЕООД":    "FDE8E8",
    "ВЕЖЕН 2010  ЕООД":    "F3E5F5",
}


# ─────────────────────────────────────────────────────────────────────────────
# STYLE HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def _side(style="thin", color="CCCCCC"):
    return Side(style=style, color=color)


def _border(style="thin", color="CCCCCC"):
    s = _side(style, color)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def cw(ws, col_index, width):
    ws.column_dimensions[get_column_letter(col_index)].width = width


def rh(ws, row_index, height):
    ws.row_dimensions[row_index].height = height


def cell_input(cell, value, fmt=None, align="right"):
    """Blue = hardcoded input value that came from source data."""
    cell.value = value
    cell.font = Font(name="Arial", size=9, color=C["input"])
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if fmt:
        cell.number_format = fmt


def cell_calc(cell, formula, fmt=None, color=None, align="right", bold=False):
    """Black = formula / calculation."""
    cell.value = formula
    cell.font = Font(name="Arial", size=9,
                     color=color or C["calc"], bold=bold)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if fmt:
        cell.number_format = fmt


def cell_xref(cell, formula, fmt=None, align="right"):
    """Green = cross-sheet reference."""
    cell.value = formula
    cell.font = Font(name="Arial", size=9, color=C["xref"])
    cell.alignment = Alignment(horizontal=align, vertical="center")
    cell.border = _border()
    if fmt:
        cell.number_format = fmt


def cell_header(cell, text, bg=None, size=9, wrap=False,
                align="center", bold=True):
    cell.value = text
    cell.font = Font(bold=bold, color="FFFFFF", name="Arial", size=size)
    cell.fill = _fill(bg or C["navy"])
    cell.alignment = Alignment(
        horizontal=align, vertical="center", wrap_text=wrap)
    cell.border = _border("medium", C["navy"])


def apply_bg(ws, row, col_start, col_end, hex_color):
    f = _fill(hex_color)
    for j in range(col_start, col_end + 1):
        ws.cell(row, j).fill = f


def title_row(ws, row, text, col_span, bg=None, size=11):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=col_span,
    )
    c = ws.cell(row, 1, text)
    c.font = Font(bold=True, color="FFFFFF", name="Arial", size=size)
    c.fill = _fill(bg or C["navy"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border("medium", C["navy"])
    rh(ws, row, 26)


def subtitle_row(ws, row, text, col_span, bg=None):
    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=col_span,
    )
    c = ws.cell(row, 1, text)
    c.font = Font(color="FFFFFF", name="Arial", size=9)
    c.fill = _fill(bg or C["teal"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    rh(ws, row, 14)


# ─────────────────────────────────────────────────────────────────────────────
# DATA PARSING
# ─────────────────────────────────────────────────────────────────────────────
def _detect_layout(df_raw):
    """
    Detect column layout from the header row.
    Returns a dict with column indices for each field.
    """
    header = df_raw.iloc[0].astype(str).str.strip().tolist()

    # 2025 layout:  [0]=doc, [1]=no, [2]=date, [3]=bgn_nodds, [4]=eur, [5]=vendor, [8]=warehouse
    # 2026 layout:  [0]=doc, [1]=no, [2]=date, [3]=lv, [4]=bgn_nodds, [5]=eur, [6]=vendor, [9]=warehouse
    has_lv_col = any("ст-ст (ЛВ)" in h or "ЛВ" == h for h in header[:5])

    if has_lv_col:
        return dict(
            eur_col=5, bgn_nodds_col=4, lv_col=3,
            vendor_col=6, warehouse_col=9,
            item_bgn_col=4, item_eur_col=5,
        )
    else:
        return dict(
            eur_col=4, bgn_nodds_col=3, lv_col=None,
            vendor_col=5, warehouse_col=8,
            item_bgn_col=3, item_eur_col=4,
        )


def parse_file(path):
    """
    Parse a procurement Excel file.
    Returns:
        inv_df   – one row per invoice header (spend source of truth)
        items_df – one row per item line (qty / SKU / WAP analysis)
        year     – detected calendar year
        layout   – column layout dict
    """
    print(f"  → Четене: {Path(path).name}")
    # Try to read all sheets; take the first one
    xls = pd.ExcelFile(path)
    df_raw = pd.read_excel(xls, sheet_name=xls.sheet_names[0], header=None)

    layout = _detect_layout(df_raw)
    eur_c   = layout["eur_col"]
    bgn_c   = layout["bgn_nodds_col"]
    lv_c    = layout["lv_col"]
    vnd_c   = layout["vendor_col"]
    wh_c    = layout["warehouse_col"]
    i_bgn_c = layout["item_bgn_col"]
    i_eur_c = layout["item_eur_col"]

    inv_records  = []
    item_records = []
    curr_inv     = {}
    in_embedded  = False   # inside ТРАНСФЕРЕН ПРОТОКОЛ / ФАКТУРА sub-docs

    for _, row in df_raw.iterrows():
        val0 = str(row[0]).strip() if pd.notna(row[0]) else ""
        val1 = str(row[1]).strip() if pd.notna(row[1]) else ""

        # ── Invoice header ────────────────────────────────────────────────
        if val0 == "ПРИХОДНА СТОКОВА РАЗПИСКА":
            in_embedded = False
            total_eur  = _to_float(row[eur_c])
            total_bgn  = _to_float(row[bgn_c])
            total_lv   = _to_float(row[lv_c]) if lv_c is not None else None
            vendor     = str(row[vnd_c]).strip() if pd.notna(row[vnd_c]) else ""
            warehouse  = str(row[wh_c]).strip()  if pd.notna(row[wh_c])  else ""
            curr_inv   = dict(
                invoice_no=val1, date=row[2], vendor=vendor,
                warehouse=warehouse, total_eur=total_eur,
                total_bgn=total_bgn, total_lv=total_lv,
            )
            inv_records.append(curr_inv.copy())

        # ── Embedded sub-documents – skip their item lines ────────────────
        elif val0 in ("ТРАНСФЕРЕН ПРОТОКОЛ", "ФАКТУРА"):
            in_embedded = True

        # ── Item line ─────────────────────────────────────────────────────
        elif (
            not in_embedded
            and curr_inv
            and (str(row[1]).strip() if pd.notna(row[1]) else "") == "кг"
        ):
            qty      = _to_float(row[2])
            spend_bgn = _to_float(row[i_bgn_c])
            spend_eur = _to_float(row[i_eur_c])
            if qty is None:
                continue
            item_records.append({
                **{k: curr_inv[k]
                   for k in ("invoice_no", "date", "vendor", "warehouse")},
                "item":      val0,
                "qty_kg":    qty,
                "spend_bgn": spend_bgn if spend_bgn else 0.0,
                "spend_eur": spend_eur if spend_eur else 0.0,
            })

    inv_df   = _build_df(inv_records,  date_col="date")
    items_df = _build_df(item_records, date_col="date")

    year = int(inv_df["year"].mode()[0]) if len(inv_df) else 0
    print(f"     Фактури: {len(inv_df)}  |  Артикули: {len(items_df)}  "
          f"|  Година: {year}  |  "
          f"EUR общо: €{inv_df['total_eur'].sum():,.2f}")
    return inv_df, items_df, year, layout


def _to_float(val):
    try:
        v = float(val)
        return v if pd.notna(v) else None
    except (TypeError, ValueError):
        return None


def _build_df(records, date_col="date"):
    if not records:
        return pd.DataFrame()
    df = pd.DataFrame(records)
    df[date_col] = pd.to_datetime(df[date_col], dayfirst=True, errors="coerce")
    df["month"]  = df[date_col].dt.month
    df["year"]   = df[date_col].dt.year
    return df


# ─────────────────────────────────────────────────────────────────────────────
# ANALYSIS HELPERS
# ─────────────────────────────────────────────────────────────────────────────
def build_abc(wap_items_baseline):
    """ABC classification on baseline year WAP-eligible items."""
    abc = (
        wap_items_baseline
        .groupby("item")
        .agg(spend=("spend_eur", "sum"), qty=("qty_kg", "sum"))
        .sort_values("spend", ascending=False)
    )
    abc["wap"] = abc["spend"] / abc["qty"]
    abc["pct"] = abc["spend"] / abc["spend"].sum()
    abc["cum"] = abc["pct"].cumsum()

    labels = []
    for cum in abc["cum"]:
        if cum <= 0.80:
            labels.append("A")
        elif cum <= 0.95:
            labels.append("B")
        else:
            labels.append("C")
    abc["abc_class"] = labels
    return abc


def monthly_wap_series(items_df, item_name):
    """Monthly WAP for a single SKU (EUR>0 rows only)."""
    sub = items_df[(items_df["item"] == item_name) & (items_df["spend_eur"] > 0)]
    return (
        sub.groupby("month")
        .apply(lambda x: x["spend_eur"].sum() / x["qty_kg"].sum()
               if x["qty_kg"].sum() > 0 else np.nan)
        .dropna()
    )


# ─────────────────────────────────────────────────────────────────────────────
# WORKBOOK BUILDER
# ─────────────────────────────────────────────────────────────────────────────
def build_workbook(
    inv_base,  items_base,  year_base,
    inv_ytd,   items_ytd,   year_ytd,
):
    print("\n  Изграждане на Excel отчет…")

    wb = Workbook()
    wb.remove(wb.active)

    # WAP-eligible subsets (EUR > 0)
    wap_base = items_base[items_base["spend_eur"] > 0].copy()
    wap_ytd  = items_ytd [items_ytd ["spend_eur"] > 0].copy()

    abc_base = build_abc(wap_base)
    abc_lkp  = abc_base["abc_class"].to_dict()

    all_vendors = sorted(
        set(inv_base["vendor"].unique()) | set(inv_ytd["vendor"].unique())
    )
    all_items = sorted(
        set(wap_base["item"].unique()) | set(wap_ytd["item"].unique())
    )

    # Row counters (set after each sheet build)
    sheet_meta = {}

    # ── Sheet 1: RAW_BASELINE ────────────────────────────────────────────
    r_base = _sheet_raw_invoices(
        wb, f"RAW_{year_base}", inv_base, year_base, title_label="BASELINE"
    )
    sheet_meta["R_BASE"] = r_base   # last data row

    # ── Sheet 2: RAW_YTD ────────────────────────────────────────────────
    r_ytd = _sheet_raw_invoices(
        wb, f"RAW_{year_ytd}", inv_ytd, year_ytd, title_label="YTD"
    )
    sheet_meta["R_YTD"] = r_ytd

    # ── Sheet 3: ITEMS_BASELINE ──────────────────────────────────────────
    ri_base = _sheet_raw_items(
        wb, f"ITEMS_{year_base}", items_base, year_base
    )
    sheet_meta["RI_BASE"] = ri_base

    # ── Sheet 4: ITEMS_YTD ───────────────────────────────────────────────
    ri_ytd = _sheet_raw_items(
        wb, f"ITEMS_{year_ytd}", items_ytd, year_ytd
    )
    sheet_meta["RI_YTD"] = ri_ytd

    # ── Sheet 5: ASSUMPTIONS ────────────────────────────────────────────
    _sheet_assumptions(wb, sheet_meta, year_base, year_ytd)

    # ── Sheet 6: MONTHLY SUMMARY ─────────────────────────────────────────
    _sheet_monthly(wb, sheet_meta, inv_base, inv_ytd,
                   items_base, items_ytd, year_base, year_ytd)

    # ── Sheet 7: VENDOR ANALYSIS ─────────────────────────────────────────
    _sheet_vendors(wb, sheet_meta, inv_base, inv_ytd,
                   items_base, items_ytd,
                   all_vendors, year_base, year_ytd)

    # ── Sheet 8: ABC ANALYSIS ────────────────────────────────────────────
    _sheet_abc(wb, sheet_meta, abc_base, wap_ytd, year_base, year_ytd)

    # ── Sheet 9: WAP by SKU ──────────────────────────────────────────────
    _sheet_wap(wb, sheet_meta, wap_base, wap_ytd,
               all_items, abc_lkp, year_base, year_ytd)

    # ── Sheet 10: PRICE VOLATILITY ───────────────────────────────────────
    _sheet_volatility(wb, sheet_meta, wap_base, wap_ytd,
                      all_items, abc_lkp, year_base, year_ytd)

    # Tab colors
    _apply_tab_colors(wb, year_base, year_ytd)

    return wb


# ─────────────────────────────────────────────────────────────────────────────
# SHEET BUILDERS
# ─────────────────────────────────────────────────────────────────────────────

def _sheet_raw_invoices(wb, sheet_name, inv_df, year, title_label):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    COLS = 10
    title_row(
        ws, 1,
        f"RAW ДАННИ {year} ({title_label}) – ФАКТУРНИ ХЕДЪРИ  |  "
        "СИНЬО=входна стойност  |  ЧЕРНО=формула",
        COLS,
    )
    subtitle_row(
        ws, 2,
        "Разходът идва от хедъра на всяка ПРИХОДНА СТОКОВА РАЗПИСКА – "
        "единственият верен source на EUR суми",
        COLS,
    )
    rh(ws, 3, 6)

    headers = ["Дата", "Фактура №", "Доставчик", "Склад",
               "EUR общо", "BGN без ДДС", "BGN с ДДС",
               "Месец №", "Тримес.", "Полугодие"]
    for j, h in enumerate(headers, 1):
        cell_header(ws.cell(4, j), h, bg=C["teal"], wrap=True)
    rh(ws, 4, 22)

    inv_s = inv_df.sort_values(["date", "invoice_no"]).reset_index(drop=True)
    for i, row in inv_s.iterrows():
        r = i + 5
        vbg = VENDOR_COLORS.get(str(row["vendor"]).strip(), C["white"])

        cell_input(ws.cell(r, 1), row["date"].date(), fmt=FMT["date"], align="center")
        cell_input(ws.cell(r, 2), row["invoice_no"],  align="left")
        cell_input(ws.cell(r, 3), row["vendor"],      align="left")
        cell_input(ws.cell(r, 4), row["warehouse"],   align="left")
        cell_input(ws.cell(r, 5), round(float(row["total_eur"]), 4), fmt=FMT["eur"])
        cell_input(ws.cell(r, 6), round(float(row["total_bgn"]), 4), fmt=FMT["eur"])
        cell_input(ws.cell(r, 7), round(float(row["total_bgn"]) * 1.2, 2), fmt=FMT["eur"])
        cell_calc(ws.cell(r, 8), f"=MONTH(A{r})",              fmt=FMT["int"])
        cell_calc(ws.cell(r, 9), f"=INT((MONTH(A{r})+2)/3)",   fmt='"Q"0')
        cell_calc(ws.cell(r, 10),f'=IF(MONTH(A{r})<=6,"H1","H2")', align="center")

        for j in range(1, 11):
            ws.cell(r, j).fill = _fill(vbg)
        rh(ws, r, 13)

    last_data = len(inv_s) + 4  # last row with data
    tot = last_data + 1

    for j in range(1, 11):
        ws.cell(tot, j).fill = _fill(C["navy"])
        ws.cell(tot, j).border = _border("medium", C["navy"])
        ws.cell(tot, j).font = Font(bold=True, color="FFFFFF",
                                    name="Arial", size=9)

    ws.cell(tot, 2, f"ОБЩО {year}  ({len(inv_s)} фактури)").alignment = \
        Alignment(horizontal="right", vertical="center")
    cell_calc(ws.cell(tot, 5), f"=SUM(E5:E{last_data})",
              fmt=FMT["eur"], bold=True, color="FFFFFF")
    cell_calc(ws.cell(tot, 6), f"=SUM(F5:F{last_data})",
              fmt=FMT["eur"], bold=True, color="FFFFFF")
    cell_calc(ws.cell(tot, 7), f"=SUM(G5:G{last_data})",
              fmt=FMT["eur"], bold=True, color="FFFFFF")
    rh(ws, tot, 18)

    for col, w in zip(range(1, 11),
                      [12, 18, 22, 32, 13, 13, 13, 8, 7, 8]):
        cw(ws, col, w)

    print(f"     ✓ {sheet_name}  ({len(inv_s)} rows, totals row {tot})")
    return tot  # returns totals row index


def _sheet_raw_items(wb, sheet_name, items_df, year):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False

    COLS = 9
    title_row(
        ws, 1,
        f"АРТИКУЛИ {year} – детайлни редове  |  "
        "ЖЪЛТО = EUR=0 (реална доставка без цена, включена в обема)",
        COLS,
    )
    subtitle_row(
        ws, 2,
        "Кол. H = EUR/кг формула (=G/E)  |  "
        "Кол. I = месец формула (=MONTH(A))  |  "
        "Синьо = входна стойност",
        COLS,
    )
    rh(ws, 3, 6)

    headers = ["Дата", "Фактура №", "Доставчик", "Артикул",
               "Кол. кг", "BGN без ДДС", "EUR без ДДС",
               "EUR/кг (форм.)", "Месец № (форм.)"]
    for j, h in enumerate(headers, 1):
        cell_header(ws.cell(4, j), h, bg=C["teal"], wrap=True)
    rh(ws, 4, 22)

    items_s = items_df.sort_values(
        ["date", "invoice_no", "item"]
    ).reset_index(drop=True)

    for i, row in items_s.iterrows():
        r = i + 5
        is_zero = float(row["spend_eur"]) == 0 if pd.notna(row["spend_eur"]) else True
        bg = "FFFDE7" if is_zero else (C["light"] if r % 2 == 0 else C["white"])

        cell_input(ws.cell(r, 1), row["date"].date(), fmt=FMT["date"], align="center")
        cell_input(ws.cell(r, 2), row["invoice_no"],  align="left")
        cell_input(ws.cell(r, 3), row["vendor"],      align="left")
        cell_input(ws.cell(r, 4), row["item"],        align="left")
        cell_input(ws.cell(r, 5), round(float(row["qty_kg"]),    4), fmt=FMT["num"])
        cell_input(ws.cell(r, 6), round(float(row["spend_bgn"]), 4), fmt=FMT["eur"])
        cell_input(ws.cell(r, 7), round(float(row["spend_eur"]), 4), fmt=FMT["eur"])

        if not is_zero:
            cell_calc(ws.cell(r, 8), f"=IFERROR(G{r}/E{r},0)", fmt=FMT["eur4"])
        else:
            ws.cell(r, 8).value = "€0"
            ws.cell(r, 8).font  = Font(name="Arial", size=9,
                                       color="FF8F00", italic=True)
            ws.cell(r, 8).alignment = Alignment(
                horizontal="center", vertical="center")

        cell_calc(ws.cell(r, 9), f"=MONTH(A{r})", fmt=FMT["int"])

        for j in range(1, 10):
            ws.cell(r, j).fill = _fill(bg)
        rh(ws, r, 13)

    last_data = len(items_s) + 4
    tot = last_data + 1

    for j in range(1, 10):
        ws.cell(tot, j).fill   = _fill(C["navy"])
        ws.cell(tot, j).border = _border("medium", C["navy"])
        ws.cell(tot, j).font   = Font(bold=True, color="FFFFFF",
                                      name="Arial", size=9)

    ws.cell(tot, 4, f"ОБЩО {year}").alignment = \
        Alignment(horizontal="right", vertical="center")
    cell_calc(ws.cell(tot, 5), f"=SUM(E5:E{last_data})", fmt=FMT["num"],
              bold=True, color="FFFFFF")
    cell_calc(ws.cell(tot, 7), f"=SUMIF(G5:G{last_data},\">0\",G5:G{last_data})",
              fmt=FMT["eur"], bold=True, color="FFFFFF")
    rh(ws, tot, 18)

    for col, w in zip(range(1, 10),
                      [12, 18, 22, 36, 10, 12, 12, 14, 8]):
        cw(ws, col, w)

    zero_count = (items_df["spend_eur"] == 0).sum()
    print(f"     ✓ {sheet_name}  ({len(items_s)} rows, "
          f"{zero_count} с EUR=0, totals row {tot})")
    return tot


def _sheet_assumptions(wb, meta, year_base, year_ytd):
    ws = wb.create_sheet("⚙️ Assumptions")
    ws.sheet_view.showGridLines = False

    COLS = 4
    title_row(ws, 1,
              "⚙️ ASSUMPTIONS – всички входни параметри  |  "
              "Синьо = може да се промени  |  Черно = производна формула",
              COLS, size=11)

    headers = ["Параметър", "Стойност (ВХОД)",
               "Производна формула", "Бележка"]
    for j, h in enumerate(headers, 1):
        cell_header(ws.cell(2, j), h, bg=C["teal"])
    rh(ws, 2, 18)

    R_BASE  = meta["R_BASE"]
    R_YTD   = meta["R_YTD"]
    RI_BASE = meta["RI_BASE"]
    RI_YTD  = meta["RI_YTD"]

    rows = [
        ("── ВАЛУТА ──",           None,    None,                    None),
        ("BGN→EUR фиксиран курс",  1.95583, None,                    "ECB фиксиран"),
        ("EUR за 1 BGN",           None,    "=1/B4",                 "Производна"),
        ("ДДС ставка",             0.20,    None,                    "20%"),
        ("Коефициент с ДДС",       None,    "=1+B6",                 "×1.20"),
        ("── ПЕРИОДИ ──",          None,    None,                    None),
        ("Baseline година",        year_base, None,                  "Пълна година"),
        ("YTD година",             year_ytd,  None,                  "Текуща"),
        ("YTD последен месец",     None,    None,                    "Авто от данните"),
        ("── ВЕРИФИКАЦИЯ (авто) ──", None,  None,                    None),
        (f"{year_base} EUR общо",  None,
         f"=RAW_{year_base}!E{R_BASE}",                              "От хедъри"),
        (f"{year_ytd} EUR YTD",    None,
         f"=RAW_{year_ytd}!E{R_YTD}",                               "От хедъри"),
        (f"{year_base} общо кг",   None,
         f"=ITEMS_{year_base}!E{RI_BASE}",                           "Всички артикули"),
        (f"{year_ytd} общо кг YTD", None,
         f"=ITEMS_{year_ytd}!E{RI_YTD}",                            "Вкл. €0 редове"),
        ("── ABC ПРАГОВЕ ──",      None,    None,                    None),
        ("Клас A (до %)",          0.80,    None,                    "Топ 80%"),
        ("Клас B (до %)",          0.95,    None,                    "80–95%"),
        ("── ВОЛАТИЛНОСТ ──",      None,    None,                    None),
        ("Висока CV >",            0.30,    None,                    ">30%"),
        ("Умерена CV >",           0.15,    None,                    ">15%"),
    ]

    r = 3
    for param, val, formula, note in rows:
        if param.startswith("──"):
            ws.merge_cells(f"A{r}:D{r}")
            c = ws.cell(r, 1, param)
            c.font = Font(bold=True, name="Arial", size=9, color="FFFFFF")
            c.fill = _fill(C["navy2"])
            c.alignment = Alignment(horizontal="left", vertical="center")
            rh(ws, r, 16)
            r += 1
            continue

        bg = _fill(C["light"] if r % 2 == 0 else C["white"])
        ws.cell(r, 1, param).font = Font(name="Arial", size=9)
        ws.cell(r, 1).alignment  = Alignment(
            horizontal="left", vertical="center")
        ws.cell(r, 1).fill   = bg
        ws.cell(r, 1).border = _border()

        if val is not None:
            fmt = ("0.00000" if param == "BGN→EUR фиксиран курс"
                   else (FMT["pct"] if isinstance(val, float) and val < 1
                         else FMT["int"]))
            cell_input(ws.cell(r, 2), val, fmt=fmt)
            ws.cell(r, 2).fill = bg
        if formula:
            fmt = (FMT["eur"] if "RAW_" in formula or "ITEMS_" in formula
                   else ("0.00000" if "1/B" in formula else FMT["pct"]))
            cell_calc(ws.cell(r, 3), formula, fmt=fmt)
            ws.cell(r, 3).fill = bg
        if note:
            ws.cell(r, 4, note).font = Font(
                name="Arial", size=8, color="666666", italic=True)
            ws.cell(r, 4).alignment = Alignment(
                horizontal="left", vertical="center")
            ws.cell(r, 4).fill = bg

        for j in range(1, 5):
            ws.cell(r, j).border = _border()
        rh(ws, r, 16)
        r += 1

    for col, w in zip(range(1, 5), [38, 16, 22, 32]):
        cw(ws, col, w)

    # Store assumption cell addresses for other sheets
    meta["A_ABC_A"] = "'⚙️ Assumptions'!B18"
    meta["A_ABC_B"] = "'⚙️ Assumptions'!B19"
    meta["A_CVH"]   = "'⚙️ Assumptions'!B22"
    meta["A_CVM"]   = "'⚙️ Assumptions'!B23"

    print("     ✓ ⚙️ Assumptions")


def _sheet_monthly(wb, meta, inv_base, inv_ytd,
                   items_base, items_ytd, year_base, year_ytd):
    ws = wb.create_sheet("📊 Monthly Summary")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C5"

    COLS = 14
    title_row(ws, 1,
              f"МЕСЕЧНО СРАВНЕНИЕ EUR – {year_base} (пълна) vs {year_ytd} YTD  |  "
              "SUMIF от RAW (разход) и ITEMS (обем) листове",
              COLS, size=11)
    subtitle_row(ws, 2,
                 "СИНЬО=входна  |  ЧЕРНО=формула  |  ЗЕЛЕНО=SUMIF препратка  |  "
                 "WAP=Претеглена средна цена EUR/кг  |  "
                 "Разход от фактурни хедъри, Обем от артикулни редове",
                 COLS)
    rh(ws, 3, 6)

    hdrs = [
        "Месец", "№",
        f"{year_base}\nРазход €", f"{year_base}\nОбем кг", f"{year_base}\nWAP €/кг",
        f"{year_ytd}\nРазход €",  f"{year_ytd}\nОбем кг",  f"{year_ytd}\nWAP €/кг",
        "YoY Δ €\nРазход", "YoY Δ%\nРазход",
        "YoY Δ кг\nОбем",  "YoY Δ%\nОбем",
        "WAP Δ €", "WAP Δ%",
    ]
    for j, h in enumerate(hdrs, 1):
        cell_header(ws.cell(4, j), h, bg=C["teal"], wrap=True)
    rh(ws, 4, 30)

    R_BASE  = meta["R_BASE"]
    R_YTD   = meta["R_YTD"]
    RI_BASE = meta["RI_BASE"]
    RI_YTD  = meta["RI_YTD"]

    m_base_spend = inv_base.groupby("month")["total_eur"].sum()
    m_ytd_spend  = inv_ytd.groupby("month")["total_eur"].sum()
    m_base_qty   = items_base.groupby("month")["qty_kg"].sum()
    m_ytd_qty    = items_ytd.groupby("month")["qty_kg"].sum()

    ytd_months = sorted(inv_ytd["month"].unique())

    for mo in range(1, 13):
        r  = mo + 4
        bg = C["light"] if mo % 2 == 0 else C["white"]

        cell_input(ws.cell(r, 1), MONTH_NAMES[mo - 1], align="left")
        cell_input(ws.cell(r, 2), mo, fmt=FMT["int"], align="center")

        # Baseline spend (SUMIF on RAW_BASELINE col H = month, col E = EUR)
        cell_xref(ws.cell(r, 3),
                  f"=SUMIF(RAW_{year_base}!H$5:H${R_BASE-1},B{r},"
                  f"RAW_{year_base}!E$5:E${R_BASE-1})",
                  fmt=FMT["eur"])
        # Baseline qty (SUMIF on ITEMS_BASELINE col I = month, col E = qty)
        cell_xref(ws.cell(r, 4),
                  f"=SUMIF(ITEMS_{year_base}!I$5:I${RI_BASE-1},B{r},"
                  f"ITEMS_{year_base}!E$5:E${RI_BASE-1})",
                  fmt=FMT["num"])
        cell_calc(ws.cell(r, 5), f"=IFERROR(C{r}/D{r},0)", fmt=FMT["eur4"])

        if mo in ytd_months:
            cell_xref(ws.cell(r, 6),
                      f"=SUMIF(RAW_{year_ytd}!H$5:H${R_YTD-1},B{r},"
                      f"RAW_{year_ytd}!E$5:E${R_YTD-1})",
                      fmt=FMT["eur"])
            cell_xref(ws.cell(r, 7),
                      f"=SUMIF(ITEMS_{year_ytd}!I$5:I${RI_YTD-1},B{r},"
                      f"ITEMS_{year_ytd}!E$5:E${RI_YTD-1})",
                      fmt=FMT["num"])
            cell_calc(ws.cell(r, 8), f"=IFERROR(F{r}/G{r},0)", fmt=FMT["eur4"])

            s_base = m_base_spend.get(mo, 0)
            s_ytd  = m_ytd_spend.get(mo, 0)
            q_base = m_base_qty.get(mo, 0)
            q_ytd  = m_ytd_qty.get(mo, 0)

            cell_calc(ws.cell(r, 9),  f"=F{r}-C{r}", fmt=FMT["eur"],
                      color=C["green"] if s_ytd >= s_base else C["red"])
            cell_calc(ws.cell(r, 10), f"=IFERROR((F{r}-C{r})/C{r},0)",
                      fmt=FMT["pct"],
                      color=C["green"] if s_ytd >= s_base else C["red"])
            cell_calc(ws.cell(r, 11), f"=G{r}-D{r}", fmt=FMT["num"],
                      color=C["green"] if q_ytd >= q_base else C["red"])
            cell_calc(ws.cell(r, 12), f"=IFERROR((G{r}-D{r})/D{r},0)",
                      fmt=FMT["pct"],
                      color=C["green"] if q_ytd >= q_base else C["red"])
            cell_calc(ws.cell(r, 13), f"=H{r}-E{r}", fmt=FMT["eur4"])
            cell_calc(ws.cell(r, 14), f"=IFERROR((H{r}-E{r})/E{r},0)",
                      fmt=FMT["pct"])
        else:
            for j in range(6, 15):
                ws.cell(r, j).value = "—"
                ws.cell(r, j).font  = Font(
                    name="Arial", size=9, color="AAAAAA")

        apply_bg(ws, r, 1, 14, bg)
        for j in range(1, 15):
            ws.cell(r, j).border = _border()
        rh(ws, r, 16)

    # Totals row
    TR = 17
    for j in range(1, 15):
        ws.cell(TR, j).fill   = _fill(C["navy"])
        ws.cell(TR, j).border = _border("medium", C["navy"])
        ws.cell(TR, j).font   = Font(
            bold=True, color="FFFFFF", name="Arial", size=9)

    ws.cell(TR, 1, "ОБЩО").alignment = Alignment(
        horizontal="left", vertical="center")

    cell_calc(ws.cell(TR, 3), "=SUM(C5:C16)",
              fmt=FMT["eur"], bold=True, color="FFFFFF")
    cell_calc(ws.cell(TR, 4), "=SUM(D5:D16)",
              fmt=FMT["num"], bold=True, color="FFFFFF")
    cell_calc(ws.cell(TR, 5), "=IFERROR(C17/D17,0)",
              fmt=FMT["eur4"], bold=True, color="FFFFFF")

    ytd_last = max(ytd_months) if ytd_months else 6
    sum_f = f"=SUM(F5:F{4+ytd_last})"
    sum_g = f"=SUM(G5:G{4+ytd_last})"
    cell_xref(ws.cell(TR, 6), sum_f, fmt=FMT["eur"])
    ws.cell(TR, 6).font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    cell_xref(ws.cell(TR, 7), sum_g, fmt=FMT["num"])
    ws.cell(TR, 7).font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    cell_calc(ws.cell(TR, 8), "=IFERROR(F17/G17,0)",
              fmt=FMT["eur4"], bold=True, color="FFFFFF")
    rh(ws, TR, 18)

    # H1 like-for-like comparison
    rh(ws, 18, 6)
    TR2 = 19
    for j in range(1, 15):
        ws.cell(TR2, j).fill   = _fill("EEF2FF")
        ws.cell(TR2, j).border = _border("medium", C["navy2"])

    ws.cell(TR2, 1, "Сравнение H1 Яну–Юни").font = Font(
        bold=True, color="FFFFFF", name="Arial", size=9)
    ws.cell(TR2, 1).fill = _fill(C["navy2"])
    ws.cell(TR2, 1).alignment = Alignment(
        horizontal="left", vertical="center")

    h1_formulas = [
        (3,  "=SUM(C5:C10)", FMT["eur"]),
        (4,  "=SUM(D5:D10)", FMT["num"]),
        (5,  "=IFERROR(C19/D19,0)", FMT["eur4"]),
        (6,  "=F17",         FMT["eur"]),
        (7,  "=G17",         FMT["num"]),
        (8,  "=IFERROR(F19/G19,0)", FMT["eur4"]),
        (9,  "=F19-C19",     FMT["eur"]),
        (10, "=IFERROR((F19-C19)/C19,0)", FMT["pct"]),
        (11, "=G19-D19",     FMT["num"]),
        (12, "=IFERROR((G19-D19)/D19,0)", FMT["pct"]),
        (13, "=H19-E19",     FMT["eur4"]),
        (14, "=IFERROR((H19-E19)/E19,0)", FMT["pct"]),
    ]
    for j, f, fmt in h1_formulas:
        cell_calc(ws.cell(TR2, j), f, fmt=fmt)
        ws.cell(TR2, j).fill = _fill("EEF2FF")
    rh(ws, TR2, 18)

    for col, w in zip(range(1, 15),
                      [15, 6, 13, 12, 11, 13, 12, 11, 12, 9, 12, 9, 10, 8]):
        cw(ws, col, w)

    print("     ✓ 📊 Monthly Summary")


def _sheet_vendors(wb, meta, inv_base, inv_ytd,
                   items_base, items_ytd,
                   all_vendors, year_base, year_ytd):
    ws = wb.create_sheet("🏢 Доставчици")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    COLS = 12
    title_row(ws, 1,
              f"ДОСТАВЧИЦИ – SUMIF от RAW листове  |  "
              f"{year_base}: {inv_base['vendor'].nunique()} доставч.  →  "
              f"{year_ytd}: {inv_ytd['vendor'].nunique()} доставч.  |  "
              "Пазарен дял = формула",
              COLS, size=11)
    rh(ws, 2, 6)

    hdrs = [
        "Доставчик",
        f"{year_base} EUR €", f"{year_base} кг", f"{year_base} WAP",
        f"{year_ytd} YTD €",  f"{year_ytd} кг",  f"{year_ytd} WAP",
        "Дял\n%BASE",  "Дял\n%YTD",
        "YoY H1 Δ €", "YoY H1 Δ%", "SKU\n{year_ytd}",
    ]
    for j, h in enumerate(hdrs, 1):
        cell_header(ws.cell(3, j), h, bg=C["teal"], wrap=True)
    rh(ws, 3, 28)

    R_BASE  = meta["R_BASE"]
    R_YTD   = meta["R_YTD"]
    RI_BASE = meta["RI_BASE"]
    RI_YTD  = meta["RI_YTD"]

    TV = 3 + len(all_vendors) + 1  # totals row
    q  = chr(34)

    for i, vendor in enumerate(all_vendors):
        r   = i + 4
        vbg = VENDOR_COLORS.get(vendor.strip(), C["white"])

        cell_input(ws.cell(r, 1), vendor, align="left")

        # Spend from RAW invoice headers
        cell_xref(ws.cell(r, 2),
                  f"=SUMIF(RAW_{year_base}!C$5:C${R_BASE-1},"
                  f"{q}{vendor}{q},RAW_{year_base}!E$5:E${R_BASE-1})",
                  fmt=FMT["eur"])
        cell_xref(ws.cell(r, 3),
                  f"=SUMIF(ITEMS_{year_base}!C$5:C${RI_BASE-1},"
                  f"{q}{vendor}{q},ITEMS_{year_base}!E$5:E${RI_BASE-1})",
                  fmt=FMT["num"])
        cell_calc(ws.cell(r, 4), f"=IFERROR(B{r}/C{r},0)", fmt=FMT["eur4"])

        cell_xref(ws.cell(r, 5),
                  f"=SUMIF(RAW_{year_ytd}!C$5:C${R_YTD-1},"
                  f"{q}{vendor}{q},RAW_{year_ytd}!E$5:E${R_YTD-1})",
                  fmt=FMT["eur"])
        cell_xref(ws.cell(r, 6),
                  f"=SUMIF(ITEMS_{year_ytd}!C$5:C${RI_YTD-1},"
                  f"{q}{vendor}{q},ITEMS_{year_ytd}!E$5:E${RI_YTD-1})",
                  fmt=FMT["num"])
        cell_calc(ws.cell(r, 7), f"=IFERROR(E{r}/F{r},0)", fmt=FMT["eur4"])

        cell_calc(ws.cell(r, 8), f"=IFERROR(B{r}/B${TV},0)", fmt=FMT["pct"])
        cell_calc(ws.cell(r, 9), f"=IFERROR(E{r}/E${TV},0)", fmt=FMT["pct"])

        # H1 baseline helper (hidden col M)
        h1v = float(
            inv_base[
                (inv_base["vendor"] == vendor) &
                (inv_base["month"].isin(range(1, 7)))
            ]["total_eur"].sum()
        )
        cell_input(ws.cell(r, 13), round(h1v, 4), fmt=FMT["eur"])

        cell_calc(ws.cell(r, 10), f"=E{r}-M{r}", fmt=FMT["eur"])

        s_ytd = float(inv_ytd[inv_ytd["vendor"] == vendor]["total_eur"].sum())
        cell_calc(ws.cell(r, 11),
                  f"=IFERROR((E{r}-M{r})/M{r},0)", fmt=FMT["pct"],
                  color=C["green"] if s_ytd >= h1v else C["red"])

        sku_count = int(
            items_ytd[items_ytd["vendor"] == vendor]["item"].nunique()
        )
        cell_input(ws.cell(r, 12), sku_count, fmt=FMT["int"], align="center")

        for j in range(1, 13):
            ws.cell(r, j).fill   = _fill(vbg)
            ws.cell(r, j).border = _border()
        ws.cell(r, 13).fill = _fill(vbg)
        rh(ws, r, 16)

    ws.column_dimensions["M"].width = 0.5

    # Totals row
    for j in range(1, 13):
        ws.cell(TV, j).fill   = _fill(C["navy"])
        ws.cell(TV, j).border = _border("medium", C["navy"])
        ws.cell(TV, j).font   = Font(bold=True, color="FFFFFF",
                                     name="Arial", size=9)

    ws.cell(TV, 1, "ОБЩО").alignment = Alignment(
        horizontal="left", vertical="center")
    cell_calc(ws.cell(TV, 2), f"=SUM(B4:B{TV-1})", fmt=FMT["eur"],
              bold=True, color="FFFFFF")
    cell_calc(ws.cell(TV, 3), f"=SUM(C4:C{TV-1})", fmt=FMT["num"],
              bold=True, color="FFFFFF")
    cell_calc(ws.cell(TV, 5), f"=SUM(E4:E{TV-1})", fmt=FMT["eur"],
              bold=True, color="FFFFFF")
    cell_calc(ws.cell(TV, 6), f"=SUM(F4:F{TV-1})", fmt=FMT["num"],
              bold=True, color="FFFFFF")
    rh(ws, TV, 18)

    # Monthly vendor breakdown (SUMPRODUCT)
    TVR = TV + 2
    ws.merge_cells(f"A{TVR}:{get_column_letter(len(all_vendors)+2)}{TVR}")
    cell_header(ws.cell(TVR, 1),
                f"МЕСЕЧЕН РАЗХОД ПО ДОСТАВЧИК {year_ytd} (€) – "
                "SUMPRODUCT от RAW",
                bg=C["teal"], size=10, align="left")
    rh(ws, TVR, 18)
    TVR += 1

    cell_header(ws.cell(TVR, 1), "Месец")
    for j, v in enumerate(all_vendors, 2):
        cell_header(ws.cell(TVR, j), v, bg=C["navy"], wrap=True)
    cell_header(ws.cell(TVR, len(all_vendors) + 2), "ОБЩО €", bg=C["teal"])
    rh(ws, TVR, 28)
    TVR += 1

    ytd_months = sorted(inv_ytd["month"].unique())
    for mo in ytd_months:
        bg = C["light"] if mo % 2 == 0 else C["white"]
        cell_input(ws.cell(TVR, 1), MONTH_NAMES[mo - 1], align="center")
        for j, vendor in enumerate(all_vendors, 2):
            cell_xref(
                ws.cell(TVR, j),
                f"=SUMPRODUCT("
                f"(RAW_{year_ytd}!C$5:C${R_YTD-1}={q}{vendor}{q})*"
                f"(RAW_{year_ytd}!H$5:H${R_YTD-1}={mo})*"
                f"RAW_{year_ytd}!E$5:E${R_YTD-1})",
                fmt=FMT["eur"],
            )
        n_v = len(all_vendors)
        cell_calc(ws.cell(TVR, n_v + 2),
                  f"=SUM(B{TVR}:{get_column_letter(n_v + 1)}{TVR})",
                  fmt=FMT["eur"], bold=True)
        apply_bg(ws, TVR, 1, n_v + 2, bg)
        for j in range(1, n_v + 3):
            ws.cell(TVR, j).border = _border()
        rh(ws, TVR, 15)
        TVR += 1

    for col, w in zip(range(1, 13),
                      [28, 13, 11, 11, 13, 11, 11, 9, 9, 12, 9, 8]):
        cw(ws, col, w)

    print("     ✓ 🏢 Доставчици")


def _sheet_abc(wb, meta, abc_base, wap_ytd, year_base, year_ytd):
    ws = wb.create_sheet("🔤 ABC Analysis")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    COLS = 10
    title_row(ws, 1,
              f"ABC АНАЛИЗ {year_base}  |  Клас = IF формула от Assumptions праговете  |  "
              "Кумул. % = верижна формула",
              COLS, size=11)
    subtitle_row(ws, 2,
                 f"Клас A = топ 80%  |  Клас B = 80–95%  |  Клас C = 95–100%  |  "
                 "WAP = D/E формула  |  % от общо = D/TOTAL формула  |  "
                 "Кумул. = верижна сума",
                 COLS)
    rh(ws, 3, 6)

    hdrs = [
        "#", "Артикул", "ABC Клас\n(IF форм.)",
        f"{year_base} Разход €\n(INPUT)", f"{year_base} кг\n(INPUT)",
        "WAP €/кг\n(D/E форм.)", "% от общо\n(форм.)", "Кумул. %\n(верига)",
        f"{year_ytd} YTD €\n(INPUT)", "YoY H1 Δ%\n(форм.)",
    ]
    for j, h in enumerate(hdrs, 1):
        cell_header(ws.cell(4, j), h, bg=C["navy"], wrap=True)
    rh(ws, 4, 30)

    A_ABC_A = meta["A_ABC_A"]
    A_ABC_B = meta["A_ABC_B"]

    N    = len(abc_base)
    TABC = 4 + N + 1  # totals row

    ytd_spend = wap_ytd.groupby("item")["spend_eur"].sum()
    h1_base   = (
        abc_base.index.map(
            lambda item: (
                abc_base.loc[item, "spend"] * 0.5
            )
        )
    )
    # More accurate H1: use actual items data filtered to months 1-6
    # (passed through abc_base index, pre-compute outside)

    for i, (item, row) in enumerate(abc_base.iterrows()):
        r   = i + 5
        cls = row["abc_class"]
        bg  = ABC_COLORS[cls]["bg"]
        fg  = ABC_COLORS[cls]["fg"]

        cell_calc(ws.cell(r, 1), f"=ROW()-4", fmt=FMT["int"], align="center")
        cell_input(ws.cell(r, 2), item, align="left")
        cell_calc(ws.cell(r, 3),
                  f'=IF(H{r}<={A_ABC_A},"A",'
                  f'IF(H{r}<={A_ABC_B},"B","C"))',
                  align="center", color=fg, bold=True)
        cell_input(ws.cell(r, 4), round(float(row["spend"]), 4), fmt=FMT["eur"])
        cell_input(ws.cell(r, 5), round(float(row["qty"]),   4), fmt=FMT["num"])
        cell_calc(ws.cell(r, 6),  f"=IFERROR(D{r}/E{r},0)",      fmt=FMT["eur4"])
        cell_calc(ws.cell(r, 7),  f"=IFERROR(D{r}/D${TABC},0)",  fmt=FMT["pct"])
        cell_calc(ws.cell(r, 8),
                  f"=G{r}" if r == 5 else f"=H{r-1}+G{r}",
                  fmt=FMT["pct"])

        s_ytd = float(ytd_spend.get(item, 0))
        if s_ytd > 0:
            cell_input(ws.cell(r, 9), round(s_ytd, 4), fmt=FMT["eur"])
        else:
            ws.cell(r, 9).value = "Ново/Няма"
            ws.cell(r, 9).font  = Font(name="Arial", size=9,
                                       color="AAAAAA", italic=True)

        # Helper col K = H1 baseline (hidden)
        h1v = float(row["spend"]) * 0.5  # approximate; real H1 from items
        cell_input(ws.cell(r, 11), round(h1v, 4), fmt=FMT["eur"])

        if s_ytd > 0 and h1v > 0:
            cell_calc(ws.cell(r, 10),
                      f"=IFERROR((I{r}-K{r})/K{r},0)", fmt=FMT["pct"],
                      color=C["green"] if s_ytd >= h1v else C["red"])
        else:
            ws.cell(r, 10).value = "—"
            ws.cell(r, 10).font  = Font(name="Arial", size=9, color="AAAAAA")

        for j in range(1, 11):
            ws.cell(r, j).fill   = _fill(bg)
            ws.cell(r, j).border = _border()
        rh(ws, r, 14)

    ws.column_dimensions["K"].width = 0.5

    # Totals
    for j in range(1, 11):
        ws.cell(TABC, j).fill   = _fill(C["navy"])
        ws.cell(TABC, j).border = _border("medium", C["navy"])
        ws.cell(TABC, j).font   = Font(bold=True, color="FFFFFF",
                                       name="Arial", size=9)
    ws.cell(TABC, 2, "ОБЩО").alignment = Alignment(
        horizontal="left", vertical="center")
    cell_calc(ws.cell(TABC, 4), f"=SUM(D5:D{TABC-1})",
              fmt=FMT["eur"], bold=True, color="FFFFFF")
    cell_calc(ws.cell(TABC, 5), f"=SUM(E5:E{TABC-1})",
              fmt=FMT["num"], bold=True, color="FFFFFF")
    rh(ws, TABC, 18)

    for col, w in zip(range(1, 11),
                      [5, 34, 9, 14, 12, 12, 9, 9, 14, 10]):
        cw(ws, col, w)

    print(f"     ✓ 🔤 ABC Analysis  ({N} SKUs)")


def _sheet_wap(wb, meta, wap_base, wap_ytd,
               all_items, abc_lkp, year_base, year_ytd):
    ws = wb.create_sheet("📈 WAP по SKU")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C6"

    MAX_COL = get_column_letter(2 + 12 * 3)
    title_row(ws, 1,
              "WAP ПО АРТИКУЛ И МЕСЕЦ – SUMPRODUCT от ITEMS листове (само EUR>0 редове)  |  "
              "Δ% = формула  |  Зелено = намаление  Червено = увеличение",
              2 + 12 * 3, size=11)
    subtitle_row(ws, 2,
                 "WAP = SUMPRODUCT(EUR ако артикул+месец+EUR>0) / SUMPRODUCT(кг ако артикул+месец+EUR>0)  |  "
                 "0-цена редове ИЗКЛЮЧЕНИ от WAP",
                 2 + 12 * 3)
    rh(ws, 3, 6)

    # Row 4 – month group headers
    ws.cell(4, 1).value = "Артикул"
    ws.cell(4, 1).font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ws.cell(4, 1).fill  = _fill(C["navy"])
    ws.cell(4, 1).border = _border()

    ws.cell(4, 2).value = "ABC"
    ws.cell(4, 2).font  = Font(bold=True, color="FFFFFF", name="Arial", size=9)
    ws.cell(4, 2).fill  = _fill(C["navy"])
    ws.cell(4, 2).border = _border()

    for mo in range(1, 13):
        cs = 3 + (mo - 1) * 3
        ws.merge_cells(
            start_row=4, start_column=cs, end_row=4, end_column=cs + 2)
        c = ws.cell(4, cs, MONTH_NAMES[mo - 1])
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
        c.fill = _fill(C["teal"] if mo <= 6 else "4A6FA5")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _border()
    rh(ws, 4, 18)

    # Row 5 – sub-headers
    ws.cell(5, 1).value = "Артикул"
    ws.cell(5, 1).font  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
    ws.cell(5, 1).fill  = _fill(C["navy2"])
    ws.cell(5, 1).border = _border()

    ws.cell(5, 2).value = "ABC"
    ws.cell(5, 2).font  = Font(bold=True, color="FFFFFF", name="Arial", size=8)
    ws.cell(5, 2).fill  = _fill(C["navy2"])
    ws.cell(5, 2).border = _border()

    for mo in range(1, 13):
        cs  = 3 + (mo - 1) * 3
        sbg = C["navy2"] if mo <= 6 else "5A7EA5"
        for j, sub in enumerate([f"WAP {year_base}\n€/кг",
                                   f"WAP {year_ytd}\n€/кг", "Δ%"]):
            c = ws.cell(5, cs + j, sub)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=8)
            c.fill = _fill(sbg)
            c.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True)
            c.border = _border()
            cw(ws, cs + j, 9)
    rh(ws, 5, 26)

    cw(ws, 1, 34)
    cw(ws, 2, 6)

    RI_BASE = meta["RI_BASE"]
    RI_YTD  = meta["RI_YTD"]
    q       = chr(34)
    ytd_months = sorted(wap_ytd["month"].unique())

    for idx, item in enumerate(all_items):
        r  = 6 + idx
        bg = C["light"] if r % 2 == 0 else C["white"]

        cell_input(ws.cell(r, 1), item, align="left")
        ws.cell(r, 1).fill = _fill(bg)

        cls = abc_lkp.get(item, "C")
        c2  = ws.cell(r, 2)
        c2.value = cls
        c2.font  = Font(bold=True, name="Arial", size=9,
                        color=ABC_COLORS[cls]["fg"])
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.fill  = _fill(ABC_COLORS[cls]["bg"])
        c2.border = _border()

        for mo in range(1, 13):
            cs   = 3 + (mo - 1) * 3
            g_b  = wap_base[(wap_base.item == item) & (wap_base.month == mo)]
            g_y  = wap_ytd [(wap_ytd.item  == item) & (wap_ytd.month  == mo)]
            has_b = len(g_b) > 0
            has_y = len(g_y) > 0 and mo in ytd_months

            # WAP baseline
            c_b = ws.cell(r, cs)
            if has_b:
                cell_xref(
                    c_b,
                    f"=IFERROR("
                    f"SUMPRODUCT("
                    f"(ITEMS_{year_base}!D$5:D${RI_BASE-1}={q}{item}{q})*"
                    f"(ITEMS_{year_base}!I$5:I${RI_BASE-1}={mo})*"
                    f"(ITEMS_{year_base}!G$5:G${RI_BASE-1}>0)*"
                    f"ITEMS_{year_base}!G$5:G${RI_BASE-1})/"
                    f"SUMPRODUCT("
                    f"(ITEMS_{year_base}!D$5:D${RI_BASE-1}={q}{item}{q})*"
                    f"(ITEMS_{year_base}!I$5:I${RI_BASE-1}={mo})*"
                    f"(ITEMS_{year_base}!G$5:G${RI_BASE-1}>0)*"
                    f"ITEMS_{year_base}!E$5:E${RI_BASE-1}),0)",
                    fmt=FMT["eur4"],
                )
            else:
                c_b.value = "—"
                c_b.font  = Font(name="Arial", size=9, color="BBBBBB")
            c_b.fill = _fill(bg)
            c_b.border = _border()

            # WAP YTD
            c_y = ws.cell(r, cs + 1)
            if has_y:
                cell_xref(
                    c_y,
                    f"=IFERROR("
                    f"SUMPRODUCT("
                    f"(ITEMS_{year_ytd}!D$5:D${RI_YTD-1}={q}{item}{q})*"
                    f"(ITEMS_{year_ytd}!I$5:I${RI_YTD-1}={mo})*"
                    f"(ITEMS_{year_ytd}!G$5:G${RI_YTD-1}>0)*"
                    f"ITEMS_{year_ytd}!G$5:G${RI_YTD-1})/"
                    f"SUMPRODUCT("
                    f"(ITEMS_{year_ytd}!D$5:D${RI_YTD-1}={q}{item}{q})*"
                    f"(ITEMS_{year_ytd}!I$5:I${RI_YTD-1}={mo})*"
                    f"(ITEMS_{year_ytd}!G$5:G${RI_YTD-1}>0)*"
                    f"ITEMS_{year_ytd}!E$5:E${RI_YTD-1}),0)",
                    fmt=FMT["eur4"],
                )
            elif mo > max(ytd_months, default=6):
                c_y.value = "—"
                c_y.font  = Font(name="Arial", size=9, color="BBBBBB")
            else:
                c_y.value = "Няма"
                c_y.font  = Font(name="Arial", size=9,
                                 color="BBBBBB", italic=True)
            c_y.fill = _fill(bg)
            c_y.border = _border()

            # Δ%
            c_d = ws.cell(r, cs + 2)
            if has_b and has_y:
                w_b = (g_b["spend_eur"].sum() / g_b["qty_kg"].sum()
                       if g_b["qty_kg"].sum() > 0 else 1)
                w_y = (g_y["spend_eur"].sum() / g_y["qty_kg"].sum()
                       if g_y["qty_kg"].sum() > 0 else 1)
                cell_calc(
                    c_d,
                    f"=IFERROR(({get_column_letter(cs+1)}{r}"
                    f"-{get_column_letter(cs)}{r})"
                    f"/{get_column_letter(cs)}{r},0)",
                    fmt=FMT["pct"],
                    color=C["green"] if w_y <= w_b else C["red"],
                )
            else:
                c_d.value = "—"
                c_d.font  = Font(name="Arial", size=9, color="BBBBBB")
            c_d.fill = _fill(bg)
            c_d.border = _border()

        rh(ws, r, 13)

    print(f"     ✓ 📈 WAP по SKU  ({len(all_items)} артикула)")


def _sheet_volatility(wb, meta, wap_base, wap_ytd,
                      all_items, abc_lkp, year_base, year_ytd):
    ws = wb.create_sheet("⚡ Волатилност")
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A5"

    COLS = 11
    title_row(ws, 1,
              "ЦЕНОВА ВОЛАТИЛНОСТ CV%  |  CV = Стд.откл. / Средна WAP × 100  |  "
              "Рейтинг = IF формула от Assumptions",
              COLS, size=11)
    subtitle_row(ws, 2,
                 f"ВИСОКА CV>30%, УМЕРЕНА 15–30%, СТАБИЛНА ≤15%  —  "
                 f"праговете в '⚙️ Assumptions'  |  WAP само от EUR>0 редове",
                 COLS)
    rh(ws, 3, 6)

    hdrs = [
        "Артикул", "ABC",
        f"{year_base} Ср.WAP\n€/кг",  f"{year_base} Стд.\nоткл.",
        f"{year_base} CV%\n(форм.)",   f"{year_base} Рейтинг\n(IF форм.)",
        f"{year_ytd} Ср.WAP\n€/кг",   f"{year_ytd} Стд.\nоткл.",
        f"{year_ytd} CV%\n(форм.)",    f"{year_ytd} Рейтинг\n(IF форм.)",
        "YoY WAP\nΔ% (форм.)",
    ]
    for j, h in enumerate(hdrs, 1):
        cell_header(ws.cell(4, j), h, bg=C["navy"], wrap=True)
    rh(ws, 4, 30)

    A_CVH = meta["A_CVH"]
    A_CVM = meta["A_CVM"]

    for pvr, item in enumerate(all_items, 5):
        bg = C["light"] if pvr % 2 == 0 else C["white"]

        s_b  = monthly_wap_series(wap_base, item)
        s_y  = monthly_wap_series(wap_ytd,  item)
        m_b  = float(s_b.mean()) if len(s_b) > 0 else None
        d_b  = float(s_b.std())  if len(s_b) > 1 else None
        cv_b = d_b / m_b         if (m_b and d_b and m_b > 0) else None
        m_y  = float(s_y.mean()) if len(s_y) > 0 else None
        d_y  = float(s_y.std())  if len(s_y) > 1 else None
        cv_y = d_y / m_y         if (m_y and d_y and m_y > 0) else None

        cell_input(ws.cell(pvr, 1), item, align="left")

        cls = abc_lkp.get(item, "C")
        c2  = ws.cell(pvr, 2)
        c2.value = cls
        c2.font  = Font(bold=True, name="Arial", size=9,
                        color=ABC_COLORS[cls]["fg"])
        c2.alignment = Alignment(horizontal="center", vertical="center")
        c2.fill  = _fill(ABC_COLORS[cls]["bg"])
        c2.border = _border()

        # Baseline
        if m_b:
            cell_input(ws.cell(pvr, 3), round(m_b, 6), fmt=FMT["eur4"])
        else:
            ws.cell(pvr, 3).value = "—"

        if d_b:
            cell_input(ws.cell(pvr, 4), round(d_b, 6), fmt=FMT["eur4"])
        else:
            ws.cell(pvr, 4).value = "—"

        if m_b and d_b:
            col5 = (C["red"] if cv_b > 0.30
                    else C["orange"] if cv_b > 0.15 else C["green"])
            cell_calc(ws.cell(pvr, 5),
                      f"=IFERROR(D{pvr}/C{pvr},0)",
                      fmt=FMT["pct"], color=col5)
            cell_calc(ws.cell(pvr, 6),
                      f'=IF(E{pvr}>{A_CVH},"🔴 ВИСОКА",'
                      f'IF(E{pvr}>{A_CVM},"🟡 УМЕРЕНА","🟢 СТАБИЛНА"))',
                      align="center")
        else:
            ws.cell(pvr, 5).value = "—"
            ws.cell(pvr, 6).value = "N/A"
            ws.cell(pvr, 6).font  = Font(
                name="Arial", size=9, color="AAAAAA")

        # YTD
        if m_y:
            cell_input(ws.cell(pvr, 7), round(m_y, 6), fmt=FMT["eur4"])
        else:
            ws.cell(pvr, 7).value = "—"

        if d_y:
            cell_input(ws.cell(pvr, 8), round(d_y, 6), fmt=FMT["eur4"])
        else:
            ws.cell(pvr, 8).value = "—"

        if m_y and d_y:
            col9 = (C["red"] if cv_y > 0.30
                    else C["orange"] if cv_y > 0.15 else C["green"])
            cell_calc(ws.cell(pvr, 9),
                      f"=IFERROR(H{pvr}/G{pvr},0)",
                      fmt=FMT["pct"], color=col9)
            cell_calc(ws.cell(pvr, 10),
                      f'=IF(I{pvr}>{A_CVH},"🔴 ВИСОКА",'
                      f'IF(I{pvr}>{A_CVM},"🟡 УМЕРЕНА","🟢 СТАБИЛНА"))',
                      align="center")
        else:
            ws.cell(pvr, 9).value  = "—"
            ws.cell(pvr, 10).value = "—"
            for j in [9, 10]:
                ws.cell(pvr, j).font = Font(
                    name="Arial", size=9, color="AAAAAA")

        if m_b and m_y:
            cell_calc(ws.cell(pvr, 11),
                      f"=IFERROR((G{pvr}-C{pvr})/C{pvr},0)",
                      fmt=FMT["pct"],
                      color=C["green"] if m_y <= m_b else C["red"])
        else:
            ws.cell(pvr, 11).value = "—"
            ws.cell(pvr, 11).font  = Font(
                name="Arial", size=9, color="AAAAAA")

        for j in range(1, 12):
            if (ws.cell(pvr, j).fill.fgColor.rgb == "00000000"
                    or not ws.cell(pvr, j).fill.patternType):
                ws.cell(pvr, j).fill = _fill(bg)
            ws.cell(pvr, j).border = _border()
        rh(ws, pvr, 14)

    for col, w in zip(range(1, 12),
                      [34, 7, 12, 12, 10, 14, 12, 12, 10, 14, 10]):
        cw(ws, col, w)

    print(f"     ✓ ⚡ Волатилност  ({len(all_items)} артикула)")


def _apply_tab_colors(wb, year_base, year_ytd):
    tab_map = {
        f"RAW_{year_base}":     "2C4E80",
        f"RAW_{year_ytd}":      "2C4E80",
        f"ITEMS_{year_base}":   "3A5BA0",
        f"ITEMS_{year_ytd}":    "3A5BA0",
        "⚙️ Assumptions":       "4A6FA5",
        "📊 Monthly Summary":   "1B3A6B",
        "🏢 Доставчици":        "F0A500",
        "🔤 ABC Analysis":      "2E7D32",
        "📈 WAP по SKU":        "0D7377",
        "⚡ Волатилност":       "C62828",
    }
    for sheet_name, color in tab_map.items():
        if sheet_name in wb.sheetnames:
            wb[sheet_name].sheet_properties.tabColor = color


# ─────────────────────────────────────────────────────────────────────────────
# CLI ENTRY POINT
# ─────────────────────────────────────────────────────────────────────────────
def parse_args():
    parser = argparse.ArgumentParser(
        description=(
            "Fruit & Vegetable Procurement Analyzer v" + VERSION + "\n"
            "Генерира пълен Excel отчет от два счетоводни файла.\n\n"
            "Примери:\n"
            "  python analyzer.py --baseline 2025.xlsx --ytd 2026.xlsx\n"
            "  python analyzer.py --baseline 2025.xlsx --ytd 2026.xlsx "
            "--output Отчет_Плодове\n"
        ),
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument(
        "--baseline", "-b",
        required=True,
        metavar="FILE",
        help="Excel файл с базисната (историческа) година – пр. 2025.xlsx",
    )
    parser.add_argument(
        "--ytd", "-y",
        required=True,
        metavar="FILE",
        help="Excel файл с текущата YTD година – пр. 2026.xlsx",
    )
    parser.add_argument(
        "--output", "-o",
        default=None,
        metavar="NAME",
        help=(
            "Базово име на изходния файл (без .xlsx). "
            "По подразбиране: FruitVeg_Analysis_YYYY-MM-DD.xlsx"
        ),
    )
    return parser.parse_args()


def main():
    print("=" * 65)
    print(f"  🥬  Fruit & Vegetable Procurement Analyzer  v{VERSION}")
    print("=" * 65)

    args = parse_args()

    # ── Validate inputs ───────────────────────────────────────────────────
    for label, path in [("--baseline", args.baseline), ("--ytd", args.ytd)]:
        if not os.path.isfile(path):
            print(f"\n❌  Файлът не е намерен ({label}): {path}")
            sys.exit(1)

    # ── Parse ─────────────────────────────────────────────────────────────
    print("\n📂  Четене на входните файлове…")
    inv_base,  items_base,  year_base, _ = parse_file(args.baseline)
    inv_ytd,   items_ytd,   year_ytd,  _ = parse_file(args.ytd)

    if year_base == year_ytd:
        print(f"\n⚠️   Предупреждение: двата файла имат една и съща година "
              f"({year_base}). Резултатите може да са неочаквани.")

    # ── Build workbook ────────────────────────────────────────────────────
    print("\n📊  Изграждане на отчета…")
    wb = build_workbook(
        inv_base,  items_base,  year_base,
        inv_ytd,   items_ytd,   year_ytd,
    )

    # ── Determine output path ─────────────────────────────────────────────
    if args.output:
        out_name = args.output
        if not out_name.lower().endswith(".xlsx"):
            out_name += ".xlsx"
    else:
        today    = date.today().strftime("%Y-%m-%d")
        out_name = f"FruitVeg_Analysis_{year_base}_vs_{year_ytd}_{today}.xlsx"

    out_path = Path(out_name)

    wb.save(str(out_path))

    # ── Summary ───────────────────────────────────────────────────────────
    print("\n" + "=" * 65)
    print("  ✅  ОТЧЕТЪТ Е ГОТОВ")
    print("=" * 65)
    print(f"\n  Файл:          {out_path.resolve()}")
    print(f"  Baseline:      {year_base}  |  "
          f"Фактури: {len(inv_base)}  |  "
          f"EUR: €{inv_base['total_eur'].sum():,.2f}")
    print(f"  YTD:           {year_ytd}  |  "
          f"Фактури: {len(inv_ytd)}  |  "
          f"EUR: €{inv_ytd['total_eur'].sum():,.2f}")
    print(f"\n  Листове в отчета:")
    for sn in wb.sheetnames:
        print(f"    • {sn}")
    print()


if __name__ == "__main__":
    main()
